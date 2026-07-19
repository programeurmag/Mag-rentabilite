"""
Génère le fichier Excel hebdomadaire (spec 1.6, point 2), avec la même
structure que le fichier de référence MAG_rentabilite_reelle.xlsx :
Lisez-moi / Taux / Heures / Attribution / Jobs / Dashboard.

Les colonnes de calcul (coûts, marges, $/h, sommes par étape/employé) sont de
VRAIES FORMULES Excel qui référencent les autres onglets, pas des valeurs
figées : si Justin corrige un taux horaire ou ajoute un montant de matériaux,
tout se recalcule automatiquement (comme dans le fichier de référence).
"""

from __future__ import annotations

from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from attribution import ResultatAttribution
from parseur import EntreeTimesheet, Job, deduire_etape
from rapport import RapportSemaine

POLICE = "Arial"

BLEU_INPUT = Font(name=POLICE, color="0000FF")
NOIR_FORMULE = Font(name=POLICE, color="000000")
GRAS = Font(name=POLICE, bold=True)
GRAS_BLANC = Font(name=POLICE, bold=True, color="FFFFFF")
REMPLISSAGE_ENTETE = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")

FORMAT_ARGENT = '#,##0.00 $;(#,##0.00 $);"-" $'
FORMAT_HEURES = "0.00"
FORMAT_PCT = "0.0%"

ETAPES = ["Remise à neuf", "Redressement", "Nettoyage", "Réfection dalles", "Sable", "Autre"]


def _entete(ws, ligne, valeurs):
    for i, v in enumerate(valeurs, start=1):
        c = ws.cell(row=ligne, column=i, value=v)
        c.font = GRAS_BLANC
        c.fill = REMPLISSAGE_ENTETE
    ws.freeze_panes = ws.cell(row=ligne + 1, column=1).coordinate


def _largeurs(ws, largeurs):
    for i, l in enumerate(largeurs, start=1):
        ws.column_dimensions[get_column_letter(i)].width = l


def _plage(feuille: str, colonne: str, derniere_ligne: int) -> str:
    """Construit une référence de plage complète, ex. Attribution!$A$2:$A$14."""
    derniere_ligne = max(derniere_ligne, 2)
    return f"{feuille}!${colonne}$2:${colonne}${derniere_ligne}"


def _feuille_lisez_moi(wb, debut: date, fin: date):
    ws = wb.create_sheet("Lisez-moi")
    ws.sheet_view.showGridLines = False
    _largeurs(ws, [110])
    lignes = [
        f"RENTABILITÉ PAR JOB / EMPLOYÉ — MAG Lavage À Pression "
        f"(données Jobber du {debut.strftime('%d %B %Y')} au {fin.strftime('%d %B %Y')})",
        "",
        "Ce fichier est généré automatiquement chaque lundi matin (GitHub Action).",
        "",
        "COMMENT LES HEURES SONT ATTRIBUÉES AUX JOBS",
        "• Punch direct : l'employé a punché directement sur le job dans Jobber -> heures assignées à ce job (le plus fiable).",
        "• Réparti : punch « General » -> réparti entre les jobs où l'employé était assigné et actifs cette journée-là, au prorata du revenu de chaque job.",
        "• Non attribué : aucun job actif ce jour-là (vente, déplacement, admin). Visible dans le Dashboard.",
        "",
        "QUALITÉ DES DONNÉES",
        "• Les entrées marquées « ⚠ timer? » (>12h, ou chronomètre encore actif) sont incluses dans les calculs mais doivent être corrigées manuellement dans Jobber si c'est une erreur.",
        "",
        "MODULE 3 — RAPPORT DE CHANTIER (Form rempli par les chauffeurs)",
        "• Onglet Jobs, colonne « Source heures » : « Rapport chantier » = heures/matériaux/overhead viennent du Form (source primaire). « Ancienne logique (Form manquant) » = aucun rapport pour ce job, l'agent retombe sur les punchs Jobber (comme avant le Module 3) — corrigez en rappelant au chauffeur de remplir le Form.",
        "• Pour un job « Rapport chantier », les colonnes Matériaux/Heures/Coût MO/Overhead sont des VALEURS calculées (pas des formules Excel) : un job peut accumuler des soumissions sur plusieurs semaines, ce que cette feuille (une semaine) ne peut pas recalculer en direct. Les colonnes Marge (avant/après overhead), $/heure et Marge % restent des formules qui référencent ces valeurs dans la même ligne.",
        "• Onglet Chantier : les soumissions brutes du Form reçues cette semaine (pour audit). Onglet Jobs en cours : les jobs multi-jours pas encore terminés, avec leur % de coûts déjà dépensés vs la valeur du job.",
        "",
        "LÉGENDE : bleu = modifiable (Taux), noir = formule (ne pas toucher), italique = info (Source heures).",
    ]
    for i, texte in enumerate(lignes, start=1):
        c = ws.cell(row=i, column=1, value=texte)
        c.font = GRAS if i == 1 else Font(name=POLICE)
        c.alignment = Alignment(wrap_text=True, vertical="top")


def _feuille_taux(wb, config: dict) -> int:
    ws = wb.create_sheet("Taux")
    _entete(ws, 1, ["Employé", "Taux horaire ($)", "Facteur charges", "Coût horaire chargé ($)", "Rôle (info)"])
    _largeurs(ws, [26, 16, 16, 22, 14])

    taux_horaires = config.get("taux_horaires", {})
    facteur = config.get("facteur_charges", 1.0)
    roles_vente = set(config.get("roles_vente", []))
    ex_employes = set(config.get("ex_employes", []))
    employes_hors_jobber = config.get("employes_hors_jobber", {}) or {}

    ligne = 2
    for employe, taux in taux_horaires.items():
        ws.cell(row=ligne, column=1, value=employe).font = Font(name=POLICE)
        c_taux = ws.cell(row=ligne, column=2, value=taux)
        c_taux.font = BLEU_INPUT
        c_facteur = ws.cell(row=ligne, column=3, value=facteur)
        c_facteur.font = BLEU_INPUT
        c_cout = ws.cell(row=ligne, column=4, value=f"=B{ligne}*C{ligne}")
        c_cout.font = NOIR_FORMULE
        c_cout.number_format = FORMAT_ARGENT
        role = "Vente" if employe in roles_vente else ("Ex-employé" if employe in ex_employes else None)
        ws.cell(row=ligne, column=5, value=role).font = Font(name=POLICE)
        ligne += 1

    # Gars sans licence Jobber (Module 3 — rapportés seulement via le Form de chantier).
    for employe, taux in employes_hors_jobber.items():
        ws.cell(row=ligne, column=1, value=employe).font = Font(name=POLICE)
        c_taux = ws.cell(row=ligne, column=2, value=taux)
        c_taux.font = BLEU_INPUT
        c_facteur = ws.cell(row=ligne, column=3, value=facteur)
        c_facteur.font = BLEU_INPUT
        c_cout = ws.cell(row=ligne, column=4, value=f"=B{ligne}*C{ligne}")
        c_cout.font = NOIR_FORMULE
        c_cout.number_format = FORMAT_ARGENT
        ws.cell(row=ligne, column=5, value="Hors Jobber").font = Font(name=POLICE)
        ligne += 1

    return ligne - 1  # dernière ligne de données


def _feuille_heures(wb, entrees: list):
    ws = wb.create_sheet("Heures")
    _entete(ws, 1, ["Date", "Employé", "Début", "Fin", "Heures", "Sur quoi (Jobber)", "Job #", "Client", "Anomalie"])
    _largeurs(ws, [12, 24, 9, 9, 9, 34, 8, 26, 12])

    entrees_triees = sorted(entrees, key=lambda e: (e.date_entree, e.employe))
    ligne = 2
    for e in entrees_triees:
        ws.cell(row=ligne, column=1, value=e.date_entree.isoformat())
        ws.cell(row=ligne, column=2, value=e.employe)
        ws.cell(row=ligne, column=5, value=round(e.heures, 2)).number_format = FORMAT_HEURES
        ws.cell(row=ligne, column=6, value=e.working_on)
        ws.cell(row=ligne, column=7, value=e.job_num)
        ws.cell(row=ligne, column=8, value=e.client)
        ws.cell(row=ligne, column=9, value="⚠ timer?" if e.anomalie else "")
        for col in (1, 2, 6, 8, 9):
            ws.cell(row=ligne, column=col).font = Font(name=POLICE)
        ligne += 1

    return ligne - 1


def _feuille_attribution(wb, resultat: ResultatAttribution, lignes_chantier: list | None = None) -> int:
    ws = wb.create_sheet("Attribution")
    _entete(ws, 1, ["Job #", "Employé", "Date", "Heures", "Source", "Anomalie", "Coût MO chargé ($)", "Revenu généré ($)"])
    _largeurs(ws, [8, 24, 12, 9, 16, 10, 18, 18])

    # Les lignes issues du Rapport de chantier (Module 3, source primaire quand
    # elle existe) sont ajoutées ici pour l'audit trail, mais l'onglet Jobs ne
    # les agrège PAS via ce SUMIF pour les jobs concernés (voir _feuille_jobs) :
    # elles peuvent couvrir plusieurs semaines, ce que cet onglet (une semaine)
    # ne peut pas représenter fidèlement en formule.
    toutes_lignes = list(resultat.lignes) + list(lignes_chantier or [])
    lignes_triees = sorted(toutes_lignes, key=lambda l: (l.job_num, l.date_entree or date.min, l.employe))
    ligne = 2
    for l in lignes_triees:
        ws.cell(row=ligne, column=1, value=l.job_num)
        ws.cell(row=ligne, column=2, value=l.employe)
        ws.cell(row=ligne, column=3, value=l.date_entree.isoformat() if l.date_entree else "")
        ws.cell(row=ligne, column=4, value=round(l.heures, 4)).number_format = FORMAT_HEURES
        ws.cell(row=ligne, column=5, value=l.source)
        ws.cell(row=ligne, column=6, value="⚠ timer?" if l.anomalie else "")
        # Coût MO chargé = heures x coût horaire chargé de l'employé (onglet Taux)
        c_cout = ws.cell(
            row=ligne,
            column=7,
            value=f'=IFERROR(D{ligne}*INDEX(Taux!$D:$D,MATCH(B{ligne},Taux!$A:$A,0)),0)',
        )
        c_cout.number_format = FORMAT_ARGENT
        # Revenu généré = heures x $/h du job (onglet Jobs, colonne M) ; 0 si job pas dans l'onglet Jobs (encore ouvert)
        c_rev = ws.cell(
            row=ligne,
            column=8,
            value=f'=IFERROR(D{ligne}*INDEX(Jobs!$M:$M,MATCH(A{ligne},Jobs!$A:$A,0)),0)',
        )
        c_rev.number_format = FORMAT_ARGENT
        for col in (1, 2, 3, 5, 6):
            ws.cell(row=ligne, column=col).font = Font(name=POLICE)
        ligne += 1

    return ligne - 1


def _feuille_jobs(wb, jobs_fermes: dict, derniere_ligne_attribution: int, jobs_chantier: dict | None = None) -> int:
    """
    Colonnes G-J (Matériaux, Heures, Coût MO, Overhead) : pour un job avec un
    Rapport de chantier (Module 3), ce sont des VALEURS calculées côté Python
    (calculer_chantier), pas des formules — un job multi-jours peut accumuler
    des soumissions sur plusieurs semaines, ce qu'une formule limitée à
    l'onglet Attribution de CETTE semaine ne peut pas capturer. Pour un job
    sans Rapport de chantier (ancienne logique), Heures/Coût MO restent des
    formules SUMIF live sur Attribution, comme avant, et Matériaux/Overhead
    valent 0 (aucune donnée matériaux sans Form, spec section 4).
    Colonnes K-N (marges, $/h, %) restent des formules qui référencent F-J
    dans la même ligne : éditer une valeur en amont recalcule la marge.
    """
    jobs_chantier = jobs_chantier or {}
    ws = wb.create_sheet("Jobs")
    _entete(
        ws, 1,
        ["Job #", "Client", "Ville", "Étape", "Date fermé", "Revenu ($)", "Matériaux ($)",
         "Heures attribuées", "Coût MO ($)", "Overhead ($)", "Marge avant overhead ($)",
         "Marge après overhead ($)", "$ / heure", "Marge %", "Source heures"],
    )
    _largeurs(ws, [8, 30, 20, 16, 12, 12, 12, 14, 12, 12, 18, 18, 10, 10, 26])

    plage_job = _plage("Attribution", "A", derniere_ligne_attribution)
    plage_heures = _plage("Attribution", "D", derniere_ligne_attribution)
    plage_cout = _plage("Attribution", "G", derniere_ligne_attribution)
    ligne = 2
    for numero, job in sorted(jobs_fermes.items(), key=lambda kv: kv[1].date_fermeture or date.min):
        jc = jobs_chantier.get(numero)

        ws.cell(row=ligne, column=1, value=numero)
        ws.cell(row=ligne, column=2, value=job.client)
        ws.cell(row=ligne, column=3, value=job.ville)
        ws.cell(row=ligne, column=4, value=deduire_etape(job.line_items))
        ws.cell(row=ligne, column=5, value=job.date_fermeture.isoformat() if job.date_fermeture else "")
        ws.cell(row=ligne, column=6, value=job.revenu_total).number_format = FORMAT_ARGENT

        if jc is not None:
            c_mat = ws.cell(row=ligne, column=7, value=round(jc.materiaux, 2))
            c_h = ws.cell(row=ligne, column=8, value=round(jc.heures_personnes, 4))
            c_c = ws.cell(row=ligne, column=9, value=round(jc.cout_mo, 2))
            c_over = ws.cell(row=ligne, column=10, value=round(jc.overhead, 2))
            source = "Rapport chantier"
        else:
            c_mat = ws.cell(row=ligne, column=7, value=0)
            c_h = ws.cell(row=ligne, column=8, value=f"=SUMIF({plage_job},A{ligne},{plage_heures})")
            c_c = ws.cell(row=ligne, column=9, value=f"=SUMIF({plage_job},A{ligne},{plage_cout})")
            c_over = ws.cell(row=ligne, column=10, value=0)
            source = "Ancienne logique (Form manquant)"
        c_mat.number_format = FORMAT_ARGENT
        c_h.number_format = FORMAT_HEURES
        c_c.number_format = FORMAT_ARGENT
        c_over.number_format = FORMAT_ARGENT

        c_marge_avant = ws.cell(row=ligne, column=11, value=f"=F{ligne}-I{ligne}-G{ligne}")
        c_marge_avant.number_format = FORMAT_ARGENT
        c_marge_apres = ws.cell(row=ligne, column=12, value=f"=K{ligne}-J{ligne}")
        c_marge_apres.number_format = FORMAT_ARGENT
        c_dh = ws.cell(row=ligne, column=13, value=f"=IFERROR(F{ligne}/H{ligne},0)")
        c_dh.number_format = FORMAT_ARGENT
        c_pct = ws.cell(row=ligne, column=14, value=f"=IFERROR(L{ligne}/F{ligne},0)")
        c_pct.number_format = FORMAT_PCT
        ws.cell(row=ligne, column=15, value=source).font = Font(name=POLICE, italic=True)

        for col in (1, 2, 3, 4, 5):
            ws.cell(row=ligne, column=col).font = Font(name=POLICE)
        ligne += 1

    return ligne - 1


def _feuille_dashboard(wb, debut: date, fin: date, derniere_ligne_jobs: int, derniere_ligne_heures: int, derniere_ligne_attribution: int, derniere_ligne_taux: int):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    _largeurs(ws, [26, 18, 16, 20, 20, 20, 22, 18, 20, 12])

    titre = ws.cell(row=1, column=1, value=f"DASHBOARD RENTABILITÉ — MAG (semaine du {debut.strftime('%d')} au {fin.strftime('%d %B %Y')})")
    titre.font = Font(name=POLICE, bold=True, size=14)

    jobs_etape = _plage("Jobs", "D", derniere_ligne_jobs)
    jobs_revenu = _plage("Jobs", "F", derniere_ligne_jobs)
    jobs_heures = _plage("Jobs", "H", derniere_ligne_jobs)
    jobs_cout = _plage("Jobs", "I", derniere_ligne_jobs)
    heures_employe = _plage("Heures", "B", derniere_ligne_heures)
    heures_h = _plage("Heures", "E", derniere_ligne_heures)
    attrib_employe = _plage("Attribution", "B", derniere_ligne_attribution)
    attrib_heures = _plage("Attribution", "D", derniere_ligne_attribution)
    attrib_cout = _plage("Attribution", "G", derniere_ligne_attribution)
    attrib_revenu = _plage("Attribution", "H", derniere_ligne_attribution)

    _entete(ws, 4, ["GLOBAL", "Revenu (jobs fermés)", "Heures punchées", "Heures attribuées (fermés)",
                    "Heures attribuées (ouverts)", "Heures non attribuées", "Coût MO attribué (tous jobs)",
                    "Coût MO (jobs fermés)", "% Main d'œuvre (fermés)", "$ / h global"])
    ws.cell(row=5, column=2, value=f"=SUM({jobs_revenu})").number_format = FORMAT_ARGENT
    ws.cell(row=5, column=3, value=f"=SUM({heures_h})").number_format = FORMAT_HEURES
    ws.cell(row=5, column=4, value=f"=SUM({jobs_heures})").number_format = FORMAT_HEURES
    ws.cell(row=5, column=5, value=f"=SUM({attrib_heures})-D5").number_format = FORMAT_HEURES
    ws.cell(row=5, column=6, value=f"=C5-SUM({attrib_heures})").number_format = FORMAT_HEURES
    ws.cell(row=5, column=7, value=f"=SUM({attrib_cout})").number_format = FORMAT_ARGENT
    ws.cell(row=5, column=8, value=f"=SUM({jobs_cout})").number_format = FORMAT_ARGENT
    ws.cell(row=5, column=9, value="=IFERROR(H5/B5,0)").number_format = FORMAT_PCT
    ws.cell(row=5, column=10, value="=IFERROR(B5/D5,0)").number_format = FORMAT_ARGENT
    for col in range(2, 11):
        ws.cell(row=5, column=col).font = NOIR_FORMULE

    ws.cell(row=7, column=1, value="PAR ÉTAPE").font = GRAS
    _entete(ws, 8, ["Étape", "Revenu", "Heures attribuées", "$ / heure", "Coût MO", "% Main d'œuvre",
                    "Marge (av. matériaux)", "Marge %"])
    ligne = 9
    for etape in ETAPES:
        ws.cell(row=ligne, column=1, value=etape).font = Font(name=POLICE)
        c_rev = ws.cell(row=ligne, column=2, value=f'=SUMIF({jobs_etape},A{ligne},{jobs_revenu})')
        c_rev.number_format = FORMAT_ARGENT
        c_h = ws.cell(row=ligne, column=3, value=f'=SUMIF({jobs_etape},A{ligne},{jobs_heures})')
        c_h.number_format = FORMAT_HEURES
        c_dh = ws.cell(row=ligne, column=4, value=f"=IFERROR(B{ligne}/C{ligne},0)")
        c_dh.number_format = FORMAT_ARGENT
        c_cout = ws.cell(row=ligne, column=5, value=f'=SUMIF({jobs_etape},A{ligne},{jobs_cout})')
        c_cout.number_format = FORMAT_ARGENT
        c_pct_mo = ws.cell(row=ligne, column=6, value=f"=IFERROR(E{ligne}/B{ligne},0)")
        c_pct_mo.number_format = FORMAT_PCT
        c_marge = ws.cell(row=ligne, column=7, value=f"=B{ligne}-E{ligne}")
        c_marge.number_format = FORMAT_ARGENT
        c_pct = ws.cell(row=ligne, column=8, value=f"=IFERROR(G{ligne}/B{ligne},0)")
        c_pct.number_format = FORMAT_PCT
        ligne += 1

    ligne += 1
    ws.cell(row=ligne, column=1, value="PAR EMPLOYÉ").font = GRAS
    ligne += 1
    _entete(ws, ligne, ["Employé", "Heures punchées", "Heures sur jobs", "Coût MO chargé", "Revenu généré", "$ / h généré"])
    ligne += 1
    for i in range(max(derniere_ligne_taux - 1, 0)):
        r = ligne + i
        r_taux = 2 + i
        ws.cell(row=r, column=1, value=f"=Taux!A{r_taux}").font = NOIR_FORMULE
        c_hp = ws.cell(row=r, column=2, value=f'=SUMIF({heures_employe},A{r},{heures_h})')
        c_hp.number_format = FORMAT_HEURES
        c_hj = ws.cell(row=r, column=3, value=f'=SUMIF({attrib_employe},A{r},{attrib_heures})')
        c_hj.number_format = FORMAT_HEURES
        c_cout = ws.cell(row=r, column=4, value=f'=SUMIF({attrib_employe},A{r},{attrib_cout})')
        c_cout.number_format = FORMAT_ARGENT
        c_rev = ws.cell(row=r, column=5, value=f'=SUMIF({attrib_employe},A{r},{attrib_revenu})')
        c_rev.number_format = FORMAT_ARGENT
        c_dh = ws.cell(row=r, column=6, value=f"=IFERROR(E{r}/C{r},0)")
        c_dh.number_format = FORMAT_ARGENT


def _feuille_chantier(wb, soumissions_semaine: list) -> int:
    """Soumissions Form brutes de LA SEMAINE du rapport (module 3), pour audit — valeurs, pas formules."""
    ws = wb.create_sheet("Chantier")
    _entete(
        ws, 1,
        ["Date", "Truck / Chef d'équipe", "Job #", "Client / adresse (Form)", "Gars présents", "Arrivée", "Départ",
         "Heures", "Nb gars", "Sacs polymère", "Litres scellant", "Bacs poussière pierre",
         "Autres matériaux", "Statut", "Notes"],
    )
    _largeurs(ws, [12, 20, 8, 26, 34, 9, 9, 9, 8, 12, 12, 16, 24, 12, 30])

    lignes_triees = sorted(soumissions_semaine, key=lambda s: (s.date_jour or date.min, s.truck))
    ligne = 2
    for s in lignes_triees:
        ws.cell(row=ligne, column=1, value=s.date_jour.isoformat() if s.date_jour else "")
        ws.cell(row=ligne, column=2, value=s.truck)
        ws.cell(row=ligne, column=3, value=s.job_num)
        ws.cell(row=ligne, column=4, value=s.client_adresse)
        ws.cell(row=ligne, column=5, value=", ".join(s.gars_presents))
        ws.cell(row=ligne, column=6, value=s.heure_arrivee.strftime("%H:%M") if s.heure_arrivee else "")
        ws.cell(row=ligne, column=7, value=s.heure_depart.strftime("%H:%M") if s.heure_depart else "")
        ws.cell(row=ligne, column=8, value=round(s.duree_heures, 2)).number_format = FORMAT_HEURES
        ws.cell(row=ligne, column=9, value=len(s.gars_presents))
        ws.cell(row=ligne, column=10, value=s.sacs_polymere)
        ws.cell(row=ligne, column=11, value=s.litres_scellant)
        ws.cell(row=ligne, column=12, value=s.bacs_poussiere_pierre)
        ws.cell(row=ligne, column=13, value=s.autres_materiaux)
        ws.cell(row=ligne, column=14, value=s.statut)
        ws.cell(row=ligne, column=15, value=s.notes)
        for col in range(1, 16):
            ws.cell(row=ligne, column=col).font = Font(name=POLICE)
        ligne += 1

    if not lignes_triees:
        ws.cell(row=2, column=1, value="Aucune soumission de rapport de chantier cette semaine.").font = Font(
            name=POLICE, italic=True
        )

    return ligne - 1


def _feuille_jobs_en_cours(wb, jobs_en_cours: list):
    """Jobs multi-jours pas encore complétés : coûts accumulés à date vs valeur du job (spec 4c)."""
    ws = wb.create_sheet("Jobs en cours")
    _entete(
        ws, 1,
        ["Job #", "Client", "Truck(s)", "Valeur du job ($)", "Coûts accumulés ($)", "% Burn",
         "Dernier statut Form", "Dernière soumission", "Alerte burn"],
    )
    _largeurs(ws, [8, 30, 20, 16, 18, 10, 18, 18, 12])

    ligne = 2
    for jc in jobs_en_cours:
        ws.cell(row=ligne, column=1, value=jc.job_num).font = Font(name=POLICE)
        ws.cell(row=ligne, column=2, value=jc.client).font = Font(name=POLICE)
        ws.cell(row=ligne, column=3, value=", ".join(jc.trucks)).font = Font(name=POLICE)
        ws.cell(row=ligne, column=4, value=round(jc.valeur, 2)).number_format = FORMAT_ARGENT
        ws.cell(row=ligne, column=5, value=round(jc.couts_accumules, 2)).number_format = FORMAT_ARGENT
        ws.cell(row=ligne, column=6, value=jc.pct_burn).number_format = FORMAT_PCT
        ws.cell(row=ligne, column=7, value=jc.statut_form_dernier or "?").font = Font(name=POLICE)
        ws.cell(
            row=ligne, column=8,
            value=jc.date_derniere_soumission.isoformat() if jc.date_derniere_soumission else "",
        ).font = Font(name=POLICE)
        c_alerte = ws.cell(row=ligne, column=9, value="🔴 oui" if jc.alerte_burn else "")
        c_alerte.font = Font(name=POLICE)
        ligne += 1

    if not jobs_en_cours:
        ws.cell(row=2, column=1, value="Aucun job en cours (multi-jours) actuellement.").font = Font(
            name=POLICE, italic=True
        )


def _feuille_alertes(wb, alertes: list):
    ws = wb.create_sheet("Alertes")
    _entete(ws, 1, ["Type", "Message"])
    _largeurs(ws, [24, 90])

    libelles = {
        "dollars_heure_bas": "$/h sous le seuil",
        "timer_oublie": "Timer oublié",
        "zero_punch_visite": "0% de punch sur visite",
        "job_zero_heure": "Job fermé à 0h",
        "non_attribue_eleve": "% non attribué élevé",
        "chantier_rapport_manquant": "Rapport de chantier manquant",
        "chantier_job_introuvable": "No de job introuvable",
        "chantier_ecart_heures": "Écart heures Form / Jobber",
        "chantier_double_truck": "Employé sur 2 trucks",
        "chantier_on_revient_sans_visite": "\"On revient\" sans visite cédulée",
        "chantier_burn_eleve": "Job en cours : burn élevé",
        "chantier_contradiction_statut": "Contradiction statut Form / Jobber",
        "chantier_materiaux_a_verifier": "Matériaux \"Autres\" à vérifier",
        "chantier_nom_non_reconnu": "Nom non reconnu (taux par défaut appliqué)",
    }
    ligne = 2
    for a in alertes:
        ws.cell(row=ligne, column=1, value=libelles.get(a.type, a.type)).font = Font(name=POLICE)
        ws.cell(row=ligne, column=2, value=a.message).font = Font(name=POLICE)
        ligne += 1
    if not alertes:
        c = ws.cell(row=2, column=1, value="✅ Aucune alerte cette semaine.")
        c.font = Font(name=POLICE)


def _feuille_analyse(wb, analyse: dict | None):
    ws = wb.create_sheet("Analyse")
    ws.sheet_view.showGridLines = False
    _largeurs(ws, [100])

    if not analyse:
        c = ws.cell(
            row=1, column=1,
            value="Analyse IA non disponible cette semaine (clé Anthropic absente ou erreur temporaire).",
        )
        c.font = Font(name=POLICE, italic=True)
        return

    ligne = 1
    c = ws.cell(row=ligne, column=1, value="CONSTATS")
    c.font = GRAS
    ligne += 1
    for constat in analyse.get("constats", []):
        c = ws.cell(row=ligne, column=1, value=f"• {constat}")
        c.font = Font(name=POLICE)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ligne += 1

    ligne += 1
    c = ws.cell(row=ligne, column=1, value="RECOMMANDATIONS")
    c.font = GRAS
    ligne += 1
    for reco in analyse.get("recommandations", []):
        c = ws.cell(row=ligne, column=1, value=f"• {reco}")
        c.font = Font(name=POLICE)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ligne += 1


def generer_excel(
    chemin_sortie: str,
    jobs_fermes: dict,
    entrees: list,
    resultat: ResultatAttribution,
    config: dict,
    debut: date,
    fin: date,
    alertes: list | None = None,
    analyse: dict | None = None,
    resultat_chantier=None,  # rapport_chantier.ResultatChantier | None (Module 3)
    jobs_en_cours: list | None = None,
    soumissions_chantier: list | None = None,  # liste brute complète (matchées + orphelines)
):
    """Génère le fichier Excel complet et le sauvegarde à chemin_sortie."""
    wb = Workbook()
    wb.remove(wb.active)

    jobs_chantier = resultat_chantier.jobs if resultat_chantier else {}
    lignes_chantier = resultat_chantier.lignes_attribution if resultat_chantier else []
    # La feuille "Chantier" (audit brut) inclut TOUTES les soumissions de la
    # semaine, y compris celles dont le no de job n'a pas matché (orphelines) :
    # c'est justement là que Justin peut repérer une faute de frappe.
    soumissions_semaine = [
        s for s in (soumissions_chantier or []) if s.date_jour and debut <= s.date_jour <= fin
    ]

    _feuille_lisez_moi(wb, debut, fin)
    derniere_taux = _feuille_taux(wb, config)
    derniere_heures = _feuille_heures(wb, entrees)
    derniere_attribution = _feuille_attribution(wb, resultat, lignes_chantier)
    derniere_jobs = _feuille_jobs(wb, jobs_fermes, derniere_attribution, jobs_chantier)
    _feuille_chantier(wb, soumissions_semaine)
    _feuille_jobs_en_cours(wb, jobs_en_cours or [])
    _feuille_dashboard(wb, debut, fin, derniere_jobs, derniere_heures, derniere_attribution, derniere_taux)
    _feuille_alertes(wb, alertes or [])
    _feuille_analyse(wb, analyse)

    wb.save(chemin_sortie)
